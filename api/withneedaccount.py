import undetected_chromedriver as uc
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from telegram import Update
from telegram.ext import Application, CommandHandler, ContextTypes
from fake_useragent import UserAgent
import time
import pandas as pd
import pickle
import os
import logging
import random
import schedule
from datetime import datetime
import openpyxl
from openpyxl.utils import get_column_letter
import asyncio
import threading

# Настройки логирования
logging.basicConfig(filename='ozon_parser.log', level=logging.INFO, 
                    format='%(asctime)s - %(levelname)s - %(message)s')

# Функция инициализации undetected-chromedriver
def init_driver(headless=True, proxy=None):
    options = uc.ChromeOptions()
    if headless:
        options.add_argument("--headless")
        options.add_argument("--no-sandbox")
        options.add_argument("--disable-dev-shm-usage")
    else:
        options.add_argument("--start-maximized")
    ua = UserAgent()
    options.add_argument(f"user-agent={ua.random}")
    if proxy:
        options.add_argument(f"--proxy-server={proxy}")
    driver = uc.Chrome(options=options)
    return driver

# Функция проверки авторизации
def check_authentication(driver):
    driver.get("https://uz.ozon.com/")
    time.sleep(random.uniform(2, 4))
    try:
        if "Кабинет" in driver.page_source:
            return True
        return False
    except Exception as e:
        logging.error(f"Ошибка проверки авторизации: {e}")
        return False

# Функция проверки ограничения доступа
def check_access_restricted(driver):
    try:
        error_message = driver.find_element(By.XPATH, "//*[contains(text(), 'Доступ ограничен')]")
        return True
    except:
        return False

# Функция загрузки cookies
def load_cookies(driver):
    if os.path.exists("cookies.pkl"):
        driver.get("https://uz.ozon.com/")
        with open("cookies.pkl", "rb") as f:
            cookies = pickle.load(f)
            for cookie in cookies:
                driver.add_cookie(cookie)
        driver.refresh()
        time.sleep(random.uniform(2, 4))
        return check_authentication(driver)
    return False

# Функция сохранения cookies
def save_cookies(driver):
    with open("cookies.pkl", "wb") as f:
        pickle.dump(driver.get_cookies(), f)

# Функция ручной авторизации
def manual_login():
    driver = init_driver(headless=False)
    try:
        driver.get("https://uz.ozon.com/")
        print("Пожалуйста, войдите в аккаунт Ozon в открывшемся браузере. После входа нажмите Enter в консоли.")
        input()
        if check_authentication(driver):
            save_cookies(driver)
            print("Cookies сохранены в cookies.pkl")
            return driver
        else:
            print("Авторизация не удалась. Проверьте логин.")
            return None
    except Exception as e:
        logging.error(f"Ошибка при ручной авторизации: {e}")
        return None
    finally:
        try:
            driver.quit()
        except Exception as e:
            logging.error(f"Ошибка при закрытии драйвера: {e}")

# Функция парсинга товаров
def parse_brand_products(driver, url):
    wait = WebDriverWait(driver, 10)
    driver.get(url)
    time.sleep(random.uniform(2, 4))
    
    if check_access_restricted(driver):
        logging.error("Доступ ограничен")
        return None
    
    products = []
    
    while True:
        for _ in range(3):
            driver.execute_script("window.scrollBy(0, 500);")
            time.sleep(random.uniform(0.5, 1.5))
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(random.uniform(2, 4))

        try:
            items = wait.until(EC.presence_of_all_elements_located((By.XPATH, "//div[contains(@class, 'tile-root')]")))
            for item in items:
                try:
                    # Наименование
                    name = item.find_element(By.XPATH, ".//span[contains(@class, 'tsBody500Medium')]").text
                    
                    # Цена по Ozon-карте
                    price = "N/A"
                    try:
                        price_elem = item.find_element(By.XPATH, ".//span[contains(@class, 'tsHeadline500Medium')]")
                        price = price_elem.text
                    except:
                        pass
                    
                    # SKU (из URL, только цифры до /?)
                    link = item.find_element(By.XPATH, ".//a[contains(@class, 'tile-clickable-element')]").get_attribute("href")
                    sku = link.split("-")[-1].split("/")[0] if "-" in link else "N/A"
                    
                    products.append({
                        "Name": name,
                        "SKU": sku,
                        "Price (Ozon Card)": price
                    })
                except Exception as e:
                    logging.error(f"Ошибка парсинга товара: {e}")
        except Exception as e:
            logging.error(f"Ошибка загрузки товаров: {e}")
            break

        try:
            next_button = driver.find_element(By.XPATH, "//a[contains(@class, 'next-page')]")
            if "disabled" in next_button.get_attribute("class"):
                break
            next_button.click()
            time.sleep(random.uniform(2, 4))
        except:
            break

    return products

# Функция сохранения в Excel с настройкой ширины столбцов
def save_to_excel(products):
    try:
        df = pd.DataFrame(products)
        excel_file = "naturalino_products.xlsx"
        df.to_excel(excel_file, index=False, engine="openpyxl")
        
        # Настройка ширины столбцов
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active
        for col in range(1, 4):  # Столбцы A, B, C
            ws.column_dimensions[get_column_letter(col)].width = 75
        wb.save(excel_file)
        return excel_file
    except Exception as e:
        logging.error(f"Ошибка сохранения в Excel: {e}")
        return None

# Функция загрузки списка пользователей
def load_users():
    if os.path.exists("users.pkl"):
        with open("users.pkl", "rb") as f:
            return pickle.load(f)
    return set()

# Функция сохранения списка пользователей
def save_users(users):
    with open("users.pkl", "wb") as f:
        pickle.dump(users, f)

# Функция парсинга и отправки
async def run_parse(context: ContextTypes.DEFAULT_TYPE):
    users = load_users()
    if not users:
        logging.info("Нет зарегистрированных пользователей")
        return

    proxy = None  # Пример: "http://123.45.67.89:8080"
    driver = init_driver(headless=True, proxy=proxy)
    try:
        if not load_cookies(driver):
            error_msg = "Сессия не найдена. Пожалуйста, выполните вход через запуск программы."
            logging.error(error_msg)
            for chat_id in users:
                await context.bot.send_message(chat_id=chat_id, text=error_msg)
            return

        brand_url = "https://uz.ozon.com/brand/naturalino-100091998"
        products = parse_brand_products(driver, brand_url)

        if products is None:
            error_msg = "Доступ ограничен. Попробуйте сменить сеть/VPN или подождать."
            logging.error(error_msg)
            for chat_id in users:
                await context.bot.send_message(chat_id=chat_id, text=error_msg)
            return

        if not products:
            error_msg = "Товары не найдены или произошла ошибка."
            logging.error(error_msg)
            for chat_id in users:
                await context.bot.send_message(chat_id=chat_id, text=error_msg)
            return

        # Сохраняем в Excel
        excel_file = save_to_excel(products)
        if not excel_file:
            error_msg = "Ошибка при сохранении в Excel."
            logging.error(error_msg)
            for chat_id in users:
                await context.bot.send_message(chat_id=chat_id, text=error_msg)
            return

        # Отправляем файл всем пользователям
        for chat_id in users:
            with open(excel_file, "rb") as f:
                await context.bot.send_document(
                    chat_id=chat_id,
                    document=f,
                    caption=f"Собрано {len(products)} товаров ({datetime.now().strftime('%Y-%m-%d %H:%M:%S')})"
                )
        logging.info(f"Файл отправлен {len(users)} пользователям")

    except Exception as e:
        logging.error(f"Ошибка в run_parse: {e}")
        if check_access_restricted(driver):
            error_msg = "Доступ ограничен. Попробуйте сменить сеть/VPN или подождать."
        elif "session" in str(e).lower() or "login" in driver.current_url.lower():
            error_msg = "Сессия истекла. Требуется повторная авторизация."
        else:
            error_msg = f"Произошла ошибка: {str(e)}"
        logging.error(error_msg)
        for chat_id in users:
            await context.bot.send_message(chat_id=chat_id, text=error_msg)
    finally:
        try:
            driver.quit()
        except Exception as e:
            logging.error(f"Ошибка при закрытии драйвера: {e}")

# Telegram команда /start
async def start_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    users = load_users()
    users.add(chat_id)
    save_users(users)
    await context.bot.send_message(chat_id=chat_id, text="Вы зарегистрированы! Получите данные парсинга каждые 2 минуты.")

# Telegram команда /parse
async def parse_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    await context.bot.send_message(chat_id=chat_id, text="Запускаю парсинг...")
    users = load_users()
    if chat_id not in users:
        users.add(chat_id)
        save_users(users)
    await run_parse(context)

# Функция для запуска schedule в отдельном потоке
def run_schedule(application):
    def job():
        application.create_task(run_parse(application))
    
    schedule.every(2).minutes.do(job)
    
    while True:
        schedule.run_pending()
        time.sleep(60)

# Основной код
def main():
    # Проверяем наличие cookies
    driver = init_driver(headless=True)
    try:
        if not load_cookies(driver):
            driver.quit()
            driver = manual_login()
            if not driver:
                print("Программа завершена из-за ошибки авторизации.")
                return
    finally:
        try:
            driver.quit()
        except Exception as e:
            logging.error(f"Ошибка при закрытии драйвера: {e}")

    # Инициализация Telegram-бота
    application = Application.builder().token("7547668298:AAFhrpD9gQROtrqLfo_UIeyYEXyC31fyWSI").build()
    
    # Добавляем команды
    application.add_handler(CommandHandler("start", start_command))
    application.add_handler(CommandHandler("parse", parse_command))
    
    # Запускаем schedule в отдельном потоке
    threading.Thread(target=run_schedule, args=(application,), daemon=True).start()

    # Запускаем polling
    print("Запускаю Telegram-бот и планировщик...")
    application.run_polling(allowed_updates=Update.ALL_TYPES, stop_signals=None)

if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("Программа завершена.")