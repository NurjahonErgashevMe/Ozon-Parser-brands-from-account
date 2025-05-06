# import undetected_chromedriver as uc
# from selenium.webdriver.common.by import By
# from selenium.webdriver.support.ui import WebDriverWait
# from selenium.webdriver.support import expected_conditions as EC
# from telegram import Update, InlineKeyboardMarkup, InlineKeyboardButton
# from telegram.ext import Application, CommandHandler, ContextTypes, CallbackQueryHandler
# import time
# import pickle
# import os
# import logging
# import asyncio
# import random
# import pandas as pd
# from datetime import datetime
# from openpyxl import load_workbook
# from telegram.error import BadRequest, TelegramError
# from dotenv import load_dotenv
# from flask import Flask, request, Response
# import json
# import threading

# # Настройки логирования
# logging.basicConfig(
#     level=logging.INFO,
#     format='%(asctime)s - %(levelname)s - %(message)s',
#     handlers=[
#         logging.FileHandler('ozon_bot.log'),
#         logging.StreamHandler()
#     ]
# )

# # Загружаем переменные окружения из .env
# load_dotenv()

# # Переменные окружения
# BOT_TOKEN = os.getenv("BOT_TOKEN")
# WEBHOOK_URL = os.getenv("WEBHOOK_URL")
# COOKIES_FILE = "ozon_cookies.pkl"  # Локальный файл для Windows
# CHECK_INTERVAL = 10
# TIMEOUT = 40
# BRAND_URL = "https://uz.ozon.com/brand/naturalino-100091998"

# # Инициализация Flask
# app = Flask(__name__)

# # Инициализация Telegram Application
# telegram_app = Application.builder().token(BOT_TOKEN).build()

# # Глобальный событийный цикл для асинхронных задач
# loop = asyncio.new_event_loop()

# # Запуск цикла в отдельном потоке
# def run_loop():
#     asyncio.set_event_loop(loop)
#     loop.run_forever()

# threading.Thread(target=run_loop, daemon=True).start()

# # Инициализация драйвера
# def init_driver(headless=True):
#     options = uc.ChromeOptions()
#     if headless:
#         options.add_argument("--headless")
#         options.add_argument("--no-sandbox")
#         options.add_argument("--disable-dev-shm-usage")
#     options.add_argument("--disable-blink-features=AutomationControlled")
#     driver = uc.Chrome(options=options)
#     return driver

# # Проверка авторизации
# def check_login(driver):
#     try:
#         driver.get("https://uz.ozon.com")
#         time.sleep(1)
#         auth_indicators = ["Мой профиль", "Кабинет", "Избранное", "Мои заказы", "Выйти"]
#         page_text = driver.page_source
#         with open('page.html', 'w', encoding='utf-8') as f:
#             f.write(page_text)
#         return any(indicator in page_text for indicator in auth_indicators)
#     except Exception as e:
#         logging.error(f"Ошибка проверки авторизации: {e}")
#         return False

# # Сохранение куков
# def save_cookies(driver):
#     try:
#         current_domain = driver.current_url.split('/')[2]
#         driver.get(f"https://{current_domain}")
#         time.sleep(1)
#         cookies = driver.get_cookies()
#         with open(COOKIES_FILE, "wb") as f:
#             pickle.dump((current_domain, cookies), f)
#         return True
#     except Exception as e:
#         logging.error(f"Ошибка сохранения куков: {e}")
#         return False

# # Загрузка куков
# def load_cookies(driver):
#     if not os.path.exists(COOKIES_FILE):
#         return False
#     try:
#         with open(COOKIES_FILE, "rb") as f:
#             domain, cookies = pickle.load(f)
#         driver.get(f"https://{domain}")
#         time.sleep(1)
#         driver.delete_all_cookies()
#         for cookie in cookies:
#             driver.add_cookie(cookie)
#         driver.refresh()
#         time.sleep(1)
#         return check_login(driver)
#     except Exception as e:
#         logging.error(f"Ошибка загрузки куков: {e}")
#         return False

# # Проверка ограничения доступа
# def check_access_restricted(driver):
#     try:
#         page_text = driver.page_source
#         return "Доступ ограничен" in page_text or "Access restricted" in page_text
#     except Exception as e:
#         logging.error(f"Ошибка проверки ограничения доступа: {e}")
#         return False

# # Парсинг товаров бренда
# def parse_brand_products(driver, url):
#     wait = WebDriverWait(driver, 10)
#     driver.get(url)
#     time.sleep(random.uniform(2, 4))
#     if check_access_restricted(driver):
#         logging.error("Доступ ограничен")
#         return None
#     products = []
#     while True:
#         for _ in range(3):
#             driver.execute_script("window.scrollBy(0, 500);")
#             time.sleep(random.uniform(0.5, 1.5))
#         driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
#         time.sleep(random.uniform(2, 4))
#         try:
#             items = wait.until(EC.presence_of_all_elements_located((By.XPATH, "//div[contains(@class, 'tile-root')]")))
#             for item in items:
#                 try:
#                     name = item.find_element(By.XPATH, ".//span[contains(@class, 'tsBody500Medium')]").text
#                     price = "N/A"
#                     try:
#                         price_elem = item.find_element(By.XPATH, ".//span[contains(@class, 'tsHeadline500Medium')]")
#                         price = price_elem.text
#                     except:
#                         pass
#                     link = item.find_element(By.XPATH, ".//a[contains(@class, 'tile-clickable-element')]").get_attribute("href")
#                     sku = link.split("-")[-1].split("/")[0] if "-" in link else "N/A"
#                     products.append({
#                         "Name": name,
#                         "SKU": sku,
#                         "Price (Ozon Card)": price
#                     })
#                 except Exception as e:
#                     logging.error(f"Ошибка парсинга товара: {e}")
#         except Exception as e:
#             logging.error(f"Ошибка загрузки товаров: {e}")
#             break
#         try:
#             next_button = driver.find_element(By.XPATH, "//a[contains(@class, 'next-page')]")
#             if "disabled" in next_button.get_attribute("class"):
#                 break
#             next_button.click()
#             time.sleep(random.uniform(2, 4))
#         except:
#             break
#     return products

# async def send_reply(update: Update, text: str):
#     try:
#         if update.message:
#             await update.message.reply_text(text)
#             logging.info("Сообщение отправлено в чат")
#         elif update.callback_query and update.callback_query.message:
#             await update.callback_query.message.reply_text(text)
#             logging.info("Сообщение отправлено через callback")
#         else:
#             logging.error("Не удалось отправить сообщение: отсутствует message или callback_query.message")
#     except TelegramError as e:
#         logging.error(f"Ошибка Telegram API при отправке сообщения: {e}")
#     except Exception as e:
#         logging.error(f"Общая ошибка при отправке сообщения: {e}")

# async def check_auth_task(update: Update, driver):
#     start_time = time.time()
#     logged_in = False
#     while time.time() - start_time < TIMEOUT:
#         try:
#             if check_login(driver):
#                 if save_cookies(driver):
#                     await send_reply(update, "✅ Вы успешно авторизованы! Куки сохранены. Теперь вы можете использовать команду /parse или кнопку 'Парсить'.")
#                 else:
#                     await send_reply(update, "✅ Вы авторизованы, но куки не сохранились. Теперь вы можете использовать команду /parse или кнопку 'Парсить'.")
#                 logged_in = True
#                 break
#         except Exception as e:
#             logging.error(f"Ошибка проверки: {e}")
#         await asyncio.sleep(CHECK_INTERVAL)
#     if not logged_in:
#         await send_reply(update, "❌ Вы не вошли в систему. Попробуйте снова.")

# async def callback_query_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
#     query = update.callback_query
#     await query.answer()
#     logging.info(f"Обработка callback: {query.data}")
#     if query.data == "open_ozon":
#         try:
#             await query.message.reply_text(
#                 "🔑 Перейдите по ссылке и войдите в Ozon:\n"
#                 "👉 https://uz.ozon.com\n\n"
#                 "⏳ Проверяю авторизацию..."
#             )
#         except BadRequest as e:
#             logging.error(f"Ошибка отправки сообщения авторизации: {e}")
#             await query.message.reply_text("❌ Ошибка. Пожалуйста, попробуйте снова.")
#             return
#         driver = init_driver(headless=True)
#         try:
#             await check_auth_task(update, driver)
#         finally:
#             try:
#                 driver.quit()
#             except:
#                 pass
#     elif query.data == "parse":
#         await parse_command(update, context)

# async def start_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
#     logging.info("start_command вызван")
#     try:
#         checking_msg = await update.message.reply_text("⏳ Проверяем сессию...")
#         driver = init_driver(headless=True)
#         is_authorized = False
#         try:
#             if load_cookies(driver):
#                 is_authorized = True
#                 logging.info("Сессия загружена")
#         finally:
#             try:
#                 driver.quit()
#             except:
#                 pass
#         keyboard_buttons = [[InlineKeyboardButton("🔍 Открыть Ozon", callback_data="open_ozon")]]
#         if is_authorized:
#             keyboard_buttons.append([InlineKeyboardButton("🔄 Парсить", callback_data="parse")])
#         keyboard = InlineKeyboardMarkup(keyboard_buttons)
#         if is_authorized:
#             await checking_msg.edit_text(
#                 "✅ Вы уже авторизованы! Используйте команду /parse или кнопку 'Парсить' для парсинга товаров.",
#                 reply_markup=keyboard
#             )
#         else:
#             await checking_msg.edit_text(
#                 "👋 Нажмите кнопку, чтобы авторизоваться в Ozon.",
#                 reply_markup=keyboard
#             )
#         logging.info("Инлайн-клавиатура отправлена")
#     except TelegramError as e:
#         logging.error(f"Ошибка Telegram API в start_command: {e}")
#         await send_reply(update, "❌ Ошибка. Пожалуйста, попробуйте снова.")
#     except Exception as e:
#         logging.error(f"Общая ошибка в start_command: {e}")
#         await send_reply(update, "❌ Ошибка. Пожалуйста, попробуйте снова.")

# async def parse_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
#     if not os.path.exists(COOKIES_FILE):
#         await send_reply(update, "❌ Сессия не найдена. Пожалуйста, сначала авторизуйтесь с помощью команды /start и кнопки 'Открыть Ozon'.")
#         return
#     await send_reply(update, "⏳ Начинаю парсинг товаров бренда Naturalino...")
#     driver = init_driver(headless=True)
#     try:
#         if not load_cookies(driver):
#             await send_reply(update, "❌ Не удалось загрузить сессию. Пожалуйста, авторизуйтесь заново.")
#             return
#         products = parse_brand_products(driver, BRAND_URL)
#         if not products:
#             await send_reply(update, "❌ Не удалось получить товары. Возможно, доступ ограничен или произошла ошибка.")
#             return
#         df = pd.DataFrame(products)
#         timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
#         excel_file = f"naturalino_products_{timestamp}.xlsx"
#         df.to_excel(excel_file, index=False)
#         wb = load_workbook(excel_file)
#         ws = wb.active
#         for col in ws.columns:
#             col_letter = col[0].column_letter
#             ws.column_dimensions[col_letter].width = 75
#         wb.save(excel_file)
#         with open(excel_file, 'rb') as f:
#             if update.message:
#                 await update.message.reply_document(document=f, caption="✅ Парсинг завершен! Вот список товаров бренда Naturalino.")
#             elif update.callback_query and update.callback_query.message:
#                 await update.callback_query.message.reply_document(document=f, caption="✅ Парсинг завершен! Вот список товаров бренда Naturalino.")
#         os.remove(excel_file)
#     except Exception as e:
#         logging.error(f"Ошибка при парсинге: {e}")
#         await send_reply(update, f"❌ Произошла ошибка при парсинге: {str(e)}")
#     finally:
#         try:
#             driver.quit()
#         except:
#             pass

# # Регистрация обработчиков
# telegram_app.add_handler(CommandHandler("start", start_command))
# telegram_app.add_handler(CommandHandler("parse", parse_command))
# telegram_app.add_handler(CallbackQueryHandler(callback_query_handler))

# # Webhook обработчик
# @app.route('/webhook', methods=['POST'])
# def webhook():
#     data = request.get_json()
#     logging.info(f"Получен webhook запрос: {data}")
#     try:
#         update = Update.de_json(data, telegram_app.bot)
#         if update:
#             # Запускаем обработку в глобальном цикле
#             asyncio.run_coroutine_threadsafe(telegram_app.process_update(update), loop)
#             logging.info("Update отправлен на обработку")
#         else:
#             logging.error("Получен некорректный update")
#     except Exception as e:
#         logging.error(f"Ошибка обработки webhook: {e}")
#     return Response(status=200)

# # Инициализация приложения и установка вебхука
# async def init_app():
#     await telegram_app.initialize()
#     await telegram_app.start()
#     try:
#         await telegram_app.bot.set_webhook(url=WEBHOOK_URL)
#         logging.info(f"Webhook успешно установлен: {WEBHOOK_URL}")
#     except TelegramError as e:
#         logging.error(f"Ошибка установки webhook: {e}")
#         await asyncio.sleep(5)
#         try:
#             await telegram_app.bot.set_webhook(url=WEBHOOK_URL)
#             logging.info(f"Webhook установлен после повторной попытки: {WEBHOOK_URL}")
#         except TelegramError as e:
#             logging.error(f"Повторная ошибка установки webhook: {e}")

# # Запуск приложения
# if __name__ == "__main__":
#     logging.info(f"Запуск Flask приложения с WEBHOOK_URL={WEBHOOK_URL}")
#     asyncio.run(init_app())
#     app.run(host='0.0.0.0', port=int(os.getenv("PORT", 5000)))

import undetected_chromedriver as uc
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from telegram import Update, InlineKeyboardMarkup, InlineKeyboardButton
from telegram.ext import Application, CommandHandler, ContextTypes, CallbackQueryHandler
import time
import pickle
import os
import logging
import asyncio
import random
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from telegram.error import BadRequest, TelegramError
from dotenv import load_dotenv
from flask import Flask, request, Response
import json
import threading

# Настройки логирования
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('ozon_bot.log'),
        logging.StreamHandler()
    ]
)

# Загружаем переменные окружения из .env
load_dotenv()

# Переменные окружения
BOT_TOKEN = os.getenv("BOT_TOKEN")
WEBHOOK_URL = os.getenv("WEBHOOK_URL")
COOKIES_FILE = "ozon_cookies.pkl"
CHECK_INTERVAL = 10
TIMEOUT = 80
BRAND_URL = "https://uz.ozon.com/brand/naturalino-100091998"

# Инициализация Flask
app = Flask(__name__)

# Инициализация Telegram Application
telegram_app = Application.builder().token(BOT_TOKEN).build()

# Глобальный событийный цикл для асинхронных задач
loop = asyncio.new_event_loop()

# Запуск цикла в отдельном потоке
def run_loop():
    asyncio.set_event_loop(loop)
    loop.run_forever()

threading.Thread(target=run_loop, daemon=True).start()

# Инициализация драйвера
def init_driver(headless=True):
    options = uc.ChromeOptions()
    if headless:
        options.add_argument("--headless")
        options.add_argument("--no-sandbox")
        options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--disable-blink-features=AutomationControlled")
    driver = uc.Chrome(options=options)
    return driver

# Проверка авторизации
def check_login(driver):
    try:
        driver.get("https://uz.ozon.com")
        time.sleep(1)
        auth_indicators = ["Мой профиль", "Кабинет", "Избранное", "Мои заказы", "Выйти"]
        page_text = driver.page_source
        with open('page.html', 'w', encoding='utf-8') as f:
            f.write(page_text)
        return any(indicator in page_text for indicator in auth_indicators)
    except Exception as e:
        logging.error(f"Ошибка проверки авторизации: {e}")
        return False

# Сохранение куков
def save_cookies(driver):
    try:
        current_domain = driver.current_url.split('/')[2]
        driver.get(f"https://{current_domain}")
        time.sleep(1)
        cookies = driver.get_cookies()
        with open(COOKIES_FILE, "wb") as f:
            pickle.dump((current_domain, cookies), f)
        return True
    except Exception as e:
        logging.error(f"Ошибка сохранения куков: {e}")
        return False

# Загрузка куков
def load_cookies(driver):
    if not os.path.exists(COOKIES_FILE):
        return False
    try:
        with open(COOKIES_FILE, "rb") as f:
            domain, cookies = pickle.load(f)
        driver.get(f"https://{domain}")
        time.sleep(1)
        driver.delete_all_cookies()
        for cookie in cookies:
            driver.add_cookie(cookie)
        driver.refresh()
        time.sleep(1)
        return check_login(driver)
    except Exception as e:
        logging.error(f"Ошибка загрузки куков: {e}")
        return False

# Проверка ограничения доступа
def check_access_restricted(driver):
    try:
        page_text = driver.page_source
        return "Доступ ограничен" in page_text or "Access restricted" in page_text
    except Exception as e:
        logging.error(f"Ошибка проверки ограничения доступа: {e}")
        return False

# Парсинг товаров бренда
def parse_brand_products(driver, url):
    wait = WebDriverWait(driver, 10)
    driver.get(url)
    time.sleep(random.uniform(2, 4))
    if check_access_restricted(driver):
        logging.error("Доступ ограничен")
        return None
    products = []
    while True:
        for _ in range(3):
            driver.execute_script("window.scrollBy(0, 500);")
            time.sleep(random.uniform(0.5, 1.5))
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(random.uniform(2, 4))
        try:
            items = wait.until(EC.presence_of_all_elements_located((By.XPATH, "//div[contains(@class, 'tile-root')]")))
            for item in items:
                try:
                    name = item.find_element(By.XPATH, ".//span[contains(@class, 'tsBody500Medium')]").text
                    price = "N/A"
                    try:
                        price_elem = item.find_element(By.XPATH, ".//span[contains(@class, 'tsHeadline500Medium')]")
                        price = price_elem.text
                    except:
                        pass
                    link = item.find_element(By.XPATH, ".//a[contains(@class, 'tile-clickable-element')]").get_attribute("href")
                    sku = link.split("-")[-1].split("/")[0] if "-" in link else "N/A"
                    products.append({
                        "Name": name,
                        "SKU": sku,
                        "Price (Ozon Card)": price
                    })
                except Exception as e:
                    logging.error(f"Ошибка парсинга товара: {e}")
        except Exception as e:
            logging.error(f"Ошибка загрузки товаров: {e}")
            break
        try:
            next_button = driver.find_element(By.XPATH, "//a[contains(@class, 'next-page')]")
            if "disabled" in next_button.get_attribute("class"):
                break
            next_button.click()
            time.sleep(random.uniform(2, 4))
        except:
            break
    return products

async def send_reply(update: Update, text: str):
    try:
        if update.message:
            await update.message.reply_text(text)
            logging.info("Сообщение отправлено в чат")
        elif update.callback_query and update.callback_query.message:
            await update.callback_query.message.reply_text(text)
            logging.info("Сообщение отправлено через callback")
        else:
            logging.error("Не удалось отправить сообщение: отсутствует message или callback_query.message")
    except TelegramError as e:
        logging.error(f"Ошибка Telegram API при отправке сообщения: {e}")
    except Exception as e:
        logging.error(f"Общая ошибка при отправке сообщения: {e}")

async def check_auth_task(update: Update, driver):
    start_time = time.time()
    logged_in = False
    while time.time() - start_time < TIMEOUT:
        try:
            if check_login(driver):
                if save_cookies(driver):
                    await send_reply(update, "✅ Вы успешно авторизованы! Куки сохранены. Теперь вы можете использовать команду /parse или кнопку 'Парсить'.")
                else:
                    await send_reply(update, "✅ Вы авторизованы, но куки не сохранились. Теперь вы можете использовать команду /parse или кнопку 'Парсить'.")
                logged_in = True
                break
        except Exception as e:
            logging.error(f"Ошибка проверки: {e}")
        await asyncio.sleep(CHECK_INTERVAL)
    if not logged_in:
        await send_reply(update, "❌ Вы не вошли в систему. Попробуйте снова.")

async def callback_query_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    logging.info(f"Обработка callback: {query.data}")
    if query.data == "open_ozon":
        try:
            await query.message.reply_text(
                "🔑 Перейдите по ссылке и войдите в Ozon:\n"
                "👉 https://uz.ozon.com\n\n"
                "⏳ Проверяю авторизацию..."
            )
        except BadRequest as e:
            logging.error(f"Ошибка отправки сообщения авторизации: {e}")
            await query.message.reply_text("❌ Ошибка. Пожалуйста, попробуйте снова.")
            return
        driver = init_driver(headless=True)
        try:
            await check_auth_task(update, driver)
        finally:
            try:
                driver.quit()
            except:
                pass
    elif query.data == "parse":
        await parse_command(update, context)

async def start_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    logging.info("start_command вызван")
    try:
        checking_msg = await update.message.reply_text("⏳ Проверяем сессию...")
        driver = init_driver(headless=True)
        is_authorized = False
        try:
            if load_cookies(driver):
                is_authorized = True
                logging.info("Сессия загружена")
        finally:
            try:
                driver.quit()
            except:
                pass
        keyboard_buttons = [[InlineKeyboardButton("🔍 Открыть Ozon", callback_data="open_ozon")]]
        if is_authorized:
            keyboard_buttons.append([InlineKeyboardButton("🔄 Парсить", callback_data="parse")])
        keyboard = InlineKeyboardMarkup(keyboard_buttons)
        if is_authorized:
            await checking_msg.edit_text(
                "✅ Вы уже авторизованы! Используйте команду /parse или кнопку 'Парсить' для парсинга товаров.",
                reply_markup=keyboard
            )
        else:
            await checking_msg.edit_text(
                "👋 Нажмите кнопку, чтобы авторизоваться в Ozon.",
                reply_markup=keyboard
            )
        logging.info("Инлайн-клавиатура отправлена")
    except TelegramError as e:
        logging.error(f"Ошибка Telegram API в start_command: {e}")
        await send_reply(update, "❌ Ошибка. Пожалуйста, попробуйте снова.")
    except Exception as e:
        logging.error(f"Общая ошибка в start_command: {e}")
        await send_reply(update, "❌ Ошибка. Пожалуйста, попробуйте снова.")

async def parse_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not os.path.exists(COOKIES_FILE):
        await send_reply(update, "❌ Сессия не найдена. Пожалуйста, сначала авторизуйтесь с помощью команды /start и кнопки 'Открыть Ozon'.")
        return
    await send_reply(update, "⏳ Начинаю парсинг товаров бренда Naturalino...")
    driver = init_driver(headless=True)
    try:
        if not load_cookies(driver):
            await send_reply(update, "❌ Не удалось загрузить сессию. Пожалуйста, авторизуйтесь заново.")
            return
        products = parse_brand_products(driver, BRAND_URL)
        if not products:
            await send_reply(update, "❌ Не удалось получить товары. Возможно, доступ ограничен или произошла ошибка.")
            return
        df = pd.DataFrame(products)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_file = f"naturalino_products_{timestamp}.xlsx"
        df.to_excel(excel_file, index=False)
        wb = load_workbook(excel_file)
        ws = wb.active
        for col in ws.columns:
            col_letter = col[0].column_letter
            ws.column_dimensions[col_letter].width = 75
        wb.save(excel_file)
        with open(excel_file, 'rb') as f:
            if update.message:
                await update.message.reply_document(document=f, caption="✅ Парсинг завершен! Вот список товаров бренда Naturalino.")
            elif update.callback_query and update.callback_query.message:
                await update.callback_query.message.reply_document(document=f, caption="✅ Парсинг завершен! Вот список товаров бренда Naturalino.")
        os.remove(excel_file)
    except Exception as e:
        logging.error(f"Ошибка при парсинге: {e}")
        await send_reply(update, f"❌ Произошла ошибка при парсинге: {str(e)}")
    finally:
        try:
            driver.quit()
        except:
            pass

# Регистрация обработчиков
telegram_app.add_handler(CommandHandler("start", start_command))
telegram_app.add_handler(CommandHandler("parse", parse_command))
telegram_app.add_handler(CallbackQueryHandler(callback_query_handler))

# Webhook обработчик
@app.route('/webhook', methods=['POST'])
def webhook():
    data = request.get_json()
    logging.info(f"Получен webhook запрос: {data}")
    try:
        update = Update.de_json(data, telegram_app.bot)
        if update:
            asyncio.run_coroutine_threadsafe(telegram_app.process_update(update), loop)
            logging.info("Update отправлен на обработку")
        else:
            logging.error("Получен некорректный update")
    except Exception as e:
        logging.error(f"Ошибка обработки webhook: {e}")
    return Response(status=200)

# Инициализация приложения и установка вебхука
async def init_app():
    await telegram_app.initialize()
    await telegram_app.start()
    try:
        await telegram_app.bot.set_webhook(url=WEBHOOK_URL)
        logging.info(f"Webhook успешно установлен: {WEBHOOK_URL}")
    except TelegramError as e:
        logging.error(f"Ошибка установки webhook: {e}")
        await asyncio.sleep(5)
        try:
            await telegram_app.bot.set_webhook(url=WEBHOOK_URL)
            logging.info(f"Webhook установлен после повторной попытки: {WEBHOOK_URL}")
        except TelegramError as e:
            logging.error(f"Повторная ошибка установки webhook: {e}")

# Корневой маршрут для проверки
@app.route('/')
def index():
    return 'Flask server is running', 200

if __name__ == "__main__":
    logging.info(f"Запуск Flask приложения с WEBHOOK_URL={WEBHOOK_URL}")
    asyncio.run(init_app())
    app.run(host='0.0.0.0', port=int(os.getenv("PORT", 5000)))